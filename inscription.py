import streamlit as st
import pandas as pd
import os
import re
from datetime import datetime
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import load_workbook, Workbook
import io
import json
from PIL import Image

# Configuration de la page
st.set_page_config(
    page_title="Plateforme d'inscription - Python Géologie & Mines",
    page_icon="🐍",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Configuration Admin
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "python2025"
modules_dir = "modules_formation"
config_file = "site_config.json"
ADMIN_ONLY_PAGES = ["admin", "statistiques"]

# Initialiser les variables de session
session_defaults = {
    'admin_logged_in': False,
    'inscriptions_df': pd.DataFrame(),
    'selected_module': "Module 1",
    'show_editor': False,
    'menu_page': "accueil",
    'show_description_editor': False,
    'sidebar_collapsed': False,
    'is_mobile': False
}

for key, value in session_defaults.items():
    if key not in st.session_state:
        st.session_state[key] = value

# Liste des modules
MODULES = [
    "Module 1 - Introduction à Python",
    "Module 2 - Bases de la programmation",
    "Module 3 - Structures de données",
    "Module 4 - Fonctions et modules",
    "Module 5 - Manipulation de fichiers",
    "Module 6 - Bibliothèques géologiques",
    "Module 7 - Visualisation de données",
    "Module 8 - Projet final"
]

# Configuration par défaut du site
DEFAULT_CONFIG = {
    "site_title": "Formation Python pour Géologie & Mines",
    "site_description": """

# 🐍 **Bienvenue à la Formation Python pour les Sciences Géologiques & Minières**

## 💡 Pourquoi apprendre Python dans le domaine de la géologie et des mines ?

Python est aujourd’hui **le langage incontournable** pour l’analyse et la visualisation de données scientifiques. Dans les domaines de la géologie et des mines, il permet de :

🔹 Automatiser le traitement de données géophysiques et géochimiques
🔹 Cartographier et modéliser des structures géologiques
🔹 Simuler des processus miniers et environnementaux
🔹 Gérer et analyser des données volumineuses avec précision
🔹 Améliorer la prise de décision grâce à des visualisations interactives

**Bref, Python devient un véritable outil d’aide à la décision dans le secteur géo-minier.**



## 🎯 **Objectifs de la formation**

À la fin de cette formation, vous serez capable de :

✅ **Maîtriser les bases de Python**

* Syntaxe simple et intuitive
* Structures de données : listes, dictionnaires, tableaux
* Fonctions, boucles, conditions
* Programmation orientée objet

✅ **Appliquer Python aux problématiques géo-minières**

* Traitement et nettoyage de données issues du terrain
* Analyse statistique de données géologiques
* Visualisation de forages, profils géophysiques, cartes, etc.
* Création de modèles géologiques simplifiés

✅ **Utiliser les bibliothèques incontournables**

* **NumPy** & **Pandas** : Manipulation et analyse de données
* **Matplotlib** & **Plotly** : Graphiques et cartes interactives
* **Geopandas**, **PyGSLIB**, **lasio**, etc. : Pour les applications spécifiques en géosciences



## 👤 **À qui s’adresse cette formation ?**

Cette formation est conçue pour toute personne souhaitant intégrer le numérique et la programmation dans les métiers de la géologie et des mines :

👨‍🎓 **Étudiants** en géologie, génie minier, ou environnement
👷‍♂️ **Professionnels** du secteur minier, pétrolier ou géotechnique
🔬 **Chercheurs** en sciences de la Terre
🛠 **Ingénieurs** en exploration, production ou aménagement

*Aucun niveau avancé en programmation n’est requis. Vous apprendrez de zéro !*


## 📚 **Organisation de la formation**

📅 **Durée** : 8 modules répartis sur 4 semaines
🏫 **Format** : Présentiel ou 100% en ligne
🖥 **Prérequis** : Aisance avec l’ordinateur (Windows/Linux)
🎓 **Attestation** : Certificat délivré à la fin de la formation



## 💥 **Les plus de notre formation**

🔥 **Formation 100% adaptée au terrain géo-minier**
🔥 **Encadrement par des experts en géologie et data science**
🔥 **Exercices pratiques avec des jeux de données réels**
🔥 **Support pédagogique clair, structuré et accessible à vie**
🔥 **Accès à une communauté d’apprentissage et de collaboration**



## 📞 **Contactez-nous dès maintenant !**

📧 **Email** : [formation@tcg-expertise.com](mailto:formation@tcg-expertise.com)
📱 **Téléphone** : +226 25 45 67 67 / ‪+33779185080
🌐 **Site web** : *En construction — restez connecté !*


### 🧭 Rejoignez-nous et entrez dans le monde de la **géologie numérique avec Python**.

**➡️ Une compétence d’avenir — Une opportunité unique — Un tremplin pour votre carrière !**
Alors
*Rejoignez-nous pour une expérience d'apprentissage unique et enrichissante !*
    """,
    "site_image": None
}

def detect_mobile():
    user_agent = st.query_params.get("user_agent", [""])[0]
    mobile_keywords = ['mobile', 'android', 'iphone', 'ipad', 'windows phone']
    return any(keyword in user_agent.lower() for keyword in mobile_keywords)

def initialiser_dossier_modules():
    """Crée le dossier modules si inexistant"""
    if not os.path.exists(modules_dir):
        os.makedirs(modules_dir)
    
    for module in MODULES:
        module_file = os.path.join(modules_dir, f"{module}.txt")
        if not os.path.exists(module_file):
            with open(module_file, "w", encoding="utf-8") as f:
                f.write(f"# {module}\n\nContenu du {module.lower()} à définir...")

def initialiser_excel():
    """Crée le fichier Excel si inexistant"""
    if not os.path.exists("inscriptions.xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.title = "Inscriptions"
        ws.append(["Nom", "Prénom", "Numéro CNIB", "Téléphone", "Structure", 
                   "Période souhaitée", "Sexe", "Âge", "Niveau", "Option de suivi", "Date d'inscription"])
        wb.save("inscriptions.xlsx")

def initialiser_config():
    """Crée le fichier de configuration si inexistant"""
    if not os.path.exists(config_file):
        with open(config_file, "w", encoding="utf-8") as f:
            json.dump(DEFAULT_CONFIG, f, indent=2, ensure_ascii=False)

def charger_config():
    """Charge la configuration du site"""
    try:
        with open(config_file, "r", encoding="utf-8") as f:
            return json.load(f)
    except:
        return DEFAULT_CONFIG

def sauvegarder_config(config):
    """Sauvegarde la configuration du site"""
    with open(config_file, "w", encoding="utf-8") as f:
        json.dump(config, f, indent=2, ensure_ascii=False)

def valider_telephone(tel):
    """Valide le format du numéro de téléphone"""
    pattern = r'^(\+226|00226)?\s?[0-9]{8}$'
    return re.match(pattern, tel.replace(' ', '')) is not None

def valider_cnib(cnib):
    """Valide le format du numéro CNIB"""
    pattern = r'^[A-Z]{1,2}[0-9]{6,8}$'
    return re.match(pattern, cnib.upper()) is not None

def valider_age(age):
    """Valide l'âge (doit être entre 16 et 80 ans)"""
    try:
        age_int = int(age)
        return 16 <= age_int <= 80
    except ValueError:
        return False

def valider_nom(nom):
    """Valide le nom (pas de chiffres, minimum 2 caractères)"""
    return len(nom) >= 2 and nom.replace(' ', '').replace('-', '').isalpha()

def charger_inscriptions():
    """Charge les inscriptions depuis le fichier Excel"""
    try:
        if os.path.exists("inscriptions.xlsx"):
            df = pd.read_excel("inscriptions.xlsx")
            return df
        else:
            return pd.DataFrame()
    except Exception as e:
        st.error(f"Erreur lors du chargement des inscriptions : {str(e)}")
        return pd.DataFrame()

def sauvegarder_inscription(data):
    """Sauvegarde une nouvelle inscription"""
    try:
        wb = load_workbook("inscriptions.xlsx")
        ws = wb["Inscriptions"]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[2] == data["Numéro CNIB"]:
                return False, "Ce numéro CNIB est déjà enregistré."
        
        data_row = [
            data["Nom"], data["Prénom"], data["Numéro CNIB"], data["Téléphone"],
            data["Structure"], data["Période souhaitée"], data["Sexe"], data["Âge"],
            data["Niveau"], data["Option de suivi"], datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ]
        ws.append(data_row)
        wb.save("inscriptions.xlsx")
        return True, "Inscription enregistrée avec succès!"
    except Exception as e:
        return False, f"Erreur lors de l'enregistrement : {str(e)}"

def charger_contenu_module(module_name):
    """Charge le contenu d'un module spécifique"""
    module_file = os.path.join(modules_dir, f"{module_name}.txt")
    if os.path.exists(module_file):
        with open(module_file, "r", encoding="utf-8") as f:
            return f.read()
    return f"Veuillez cliquer sur le {module_name} pour voir le Contenu."

def sauvegarder_contenu_module(module_name, content):
    """Sauvegarde le contenu d'un module spécifique"""
    module_file = os.path.join(modules_dir, f"{module_name}.txt")
    with open(module_file, "w", encoding="utf-8") as f:
        f.write(content)

def prepare_count_data(df, column_name):
    """Prépare les données pour les graphiques à barres"""
    counts = df[column_name].value_counts().reset_index()
    counts.columns = [column_name, 'count']
    return counts

def generer_fichier_excel_download():
    """Génère un fichier Excel téléchargeable"""
    try:
        df = charger_inscriptions()
        if df.empty:
            return None
        
        buffer = io.BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Inscriptions', index=False)
            
            if not df.empty:
                stats_data = {
                    'Statistique': [
                        'Total inscriptions',
                        'Hommes',
                        'Femmes',
                        'Âge moyen',
                        'Niveau débutant',
                        'Niveau intermédiaire',
                        'Niveau avancé',
                        'Présentiel',
                        'En ligne',
                        'Hybride'
                    ],
                    'Valeur': [
                        len(df),
                        len(df[df['Sexe'] == 'Homme']),
                        len(df[df['Sexe'] == 'Femme']),
                        round(df['Âge'].mean(), 1),
                        len(df[df['Niveau'] == 'Débutant']),
                        len(df[df['Niveau'] == 'Intermédiaire']),
                        len(df[df['Niveau'] == 'Avancé']),
                        len(df[df['Option de suivi'] == 'Présentiel']),
                        len(df[df['Option de suivi'] == 'En ligne']),
                        len(df[df['Option de suivi'] == 'Hybride'])
                    ]
                }
                
                stats_df = pd.DataFrame(stats_data)
                stats_df.to_excel(writer, sheet_name='Statistiques', index=False)
        
        buffer.seek(0)
        return buffer.getvalue()
    
    except Exception as e:
        st.error(f"Erreur lors de la génération du fichier : {str(e)}")
        return None

def generer_rapport_csv():
    """Génère un rapport CSV téléchargeable"""
    try:
        df = charger_inscriptions()
        if df.empty:
            return None
        
        csv_buffer = io.StringIO()
        df.to_csv(csv_buffer, index=False, encoding='utf-8')
        csv_buffer.seek(0)
        return csv_buffer.getvalue()
    
    except Exception as e:
        st.error(f"Erreur lors de la génération du CSV : {str(e)}")
        return None

# Initialiser les dossiers et fichiers
initialiser_dossier_modules()
initialiser_excel()
initialiser_config()
st.session_state.is_mobile = detect_mobile()

# CSS personnalisé responsive
st.markdown("""
<style>
    /* Base styles */
    .main-container {
        max-width: 1200px;
        margin: 0 auto;
        padding: 1rem;
    }
    
    .card {
        background: white;
        border-radius: 12px;
        box-shadow: 0 4px 12px rgba(0,0,0,0.1);
        padding: 1.5rem;
        margin-bottom: 1.5rem;
    }
    
    .section-title {
        color: #2E86AB;
        font-size: 1.8rem;
        margin-bottom: 1rem;
        border-bottom: 2px solid #2E86AB;
        padding-bottom: 0.5rem;
    }
    
    .module-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        padding: 1.5rem;
        border-radius: 10px;
        margin-bottom: 1rem;
        cursor: pointer;
        transition: all 0.3s ease;
        border: none;
        width: 100%;
        text-align: center;
    }
    
    .module-card:hover {
        transform: translateY(-3px);
        box-shadow: 0 8px 20px rgba(0,0,0,0.2);
    }
    
    .stats-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        padding: 1rem;
        border-radius: 10px;
        margin-bottom: 1rem;
        text-align: center;
    }
    
    /* Responsive adjustments */
    @media screen and (max-width: 768px) {
        .main-container {
            padding: 0.5rem;
        }
        
        .card {
            padding: 1rem;
        }
        
        .section-title {
            font-size: 1.5rem;
        }
        
        .module-grid {
            grid-template-columns: 1fr !important;
            gap: 0.5rem;
        }
    }
    
    /* Hide streamlit branding */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
</style>
""", unsafe_allow_html=True)

# Bouton de réduction/expansion de la sidebar
if st.button("≡", key="sidebar_toggle"):
    st.session_state.sidebar_collapsed = not st.session_state.sidebar_collapsed
    st.rerun()

# SIDEBAR MENU
config = charger_config()

if not st.session_state.sidebar_collapsed:
    with st.sidebar:
        st.markdown(f"""
        <div class="card">
            <h2 style="text-align: center; color: white;">🐍 Menu Principal</h2>
        </div>
        """, unsafe_allow_html=True)
        
        # Correction appliquée ici :
        nav_options = [
            ("🏠accueil", "accueil"),
            ("📘contenu", "contenu"),
            ("📝inscription", "inscription")
        ]
        
        if st.session_state.admin_logged_in:
            nav_options.append(("📊statistiques", "statistiques"))
        
        nav_options.append(("👤admin", "admin"))
        
        for icon, page in nav_options:
            if st.button(f"{icon} {page.capitalize()}", key=f"nav_{page}", use_container_width=True):
                st.session_state.menu_page = page
                st.rerun()
        
        if st.session_state.admin_logged_in:
            st.markdown("""
            <div class="card" style="background-color: #28a74520;">
                <p style="text-align: center;">✅ Connecté en tant qu'Admin</p>
            </div>
            """, unsafe_allow_html=True)
        
        st.markdown("""
        <div class="card">
            <h4>📞 Contact</h4>
            <p>📧 formation@tcg-expertise.com</p>
            <p>📱 +226 25 45 67 67</p>
            <p>📱 ‪+33779185080</p>
        </div>
        """, unsafe_allow_html=True)

# Navigation horizontale pour mobile
if st.session_state.sidebar_collapsed:
    st.markdown('<div class="card"><h3>🧭 Navigation</h3></div>', unsafe_allow_html=True)
    
    nav_options = [
        ("🏠accueil", "accueil"),
        ("📘contenu", "contenu"),
        ("📝inscription", "inscription")
    ]
    
    if st.session_state.admin_logged_in:
        nav_options.append(("📊statistiques", "statistiques"))
    
    nav_options.append(("👤admin", "admin"))
    
    cols = st.columns(len(nav_options))
    for i, (icon, page) in enumerate(nav_options):
        with cols[i]:
            if st.button(icon, key=f"nav_mobile_{page}", use_container_width=True):
                st.session_state.menu_page = page
                st.rerun()
    
    st.markdown("---")

# CONTENU PRINCIPAL
st.markdown(f'<div class="main-container">', unsafe_allow_html=True)

# Page Administration
if st.session_state.menu_page == "admin":
    st.markdown('<div class="card"><h2 class="section-title">👤 Administration</h2></div>', unsafe_allow_html=True)
    
    if not st.session_state.admin_logged_in:
        with st.form("login_form"):
            username = st.text_input("👤 Nom d'utilisateur")
            password = st.text_input("🔒 Mot de passe", type="password")
            
            if st.form_submit_button("🚀 Se connecter", type="primary"):
                if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
                    st.session_state.admin_logged_in = True
                    st.success("✅ Connexion réussie !")
                    st.balloons()
                    st.rerun()
                else:
                    st.error("❌ Identifiants incorrects")
    else:
        st.success("✅ Vous êtes connecté en tant qu'administrateur.")
        
        if st.button("🚪 Se déconnecter", type="secondary"):
            st.session_state.admin_logged_in = False
            st.success("Déconnexion réussie.")
            st.rerun()
        
        # Téléchargement des données
        st.markdown('<div class="card"><h3>📥 Téléchargement des données</h3></div>', unsafe_allow_html=True)
        df = charger_inscriptions()
        
        if not df.empty:
            st.markdown(f"""
            <div class="card">
                <p>📈 <strong>{len(df)}</strong> inscriptions enregistrées</p>
                <p>📅 Dernière mise à jour : {datetime.now().strftime('%d/%m/%Y %H:%M')}</p>
            </div>
            """, unsafe_allow_html=True)
            
            excel_data = generer_fichier_excel_download()
            csv_data = generer_rapport_csv()
            
            if excel_data:
                st.download_button(
                    label="📊 Télécharger Excel",
                    data=excel_data,
                    file_name=f"inscriptions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            
            if csv_data:
                st.download_button(
                    label="📋 Télécharger CSV",
                    data=csv_data,
                    file_name=f"inscriptions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                    mime="text/csv"
                )
        else:
            st.warning("📭 Aucune inscription disponible")
        
        # Gestion du contenu
        st.markdown('<div class="card"><h3>✏️ Gestion du contenu</h3></div>', unsafe_allow_html=True)
        
        # Nouvelle section pour changer l'image
        st.markdown('<div class="card"><h4>🖼️ Changer l\'image du site</h4></div>', unsafe_allow_html=True)
        uploaded_image = st.file_uploader("Télécharger une nouvelle image", type=["jpg", "jpeg", "png"])
        
        if uploaded_image is not None:
            try:
                # Sauvegarder l'image
                image_path = "site_image." + uploaded_image.name.split(".")[-1]
                with open(image_path, "wb") as f:
                    f.write(uploaded_image.getbuffer())
                
                # Mettre à jour la configuration
                config["site_image"] = image_path
                sauvegarder_config(config)
                st.success("✅ Image mise à jour avec succès!")
                
                # Afficher un aperçu
                st.image(uploaded_image, caption="Nouvelle image du site", use_container_width=True)
            except Exception as e:
                st.error(f"❌ Erreur lors du téléchargement de l'image: {str(e)}")
        
        # Bouton pour supprimer l'image actuelle
        if config.get("site_image"):
            if st.button("🗑️ Supprimer l'image actuelle", type="secondary"):
                try:
                    os.remove(config["site_image"])
                    config["site_image"] = None
                    sauvegarder_config(config)
                    st.success("✅ Image supprimée avec succès!")
                except Exception as e:
                    st.error(f"❌ Erreur lors de la suppression de l'image: {str(e)}")
        
        if st.button("📝 Modifier la description du site"):
            st.session_state.show_description_editor = not st.session_state.show_description_editor
            st.rerun()
        
        if st.session_state.show_description_editor:
            new_description = st.text_area("Description du site", value=config["site_description"], height=300)
            
            if st.button("💾 Sauvegarder", type="primary"):
                config["site_description"] = new_description
                sauvegarder_config(config)
                st.success("✅ Description mise à jour !")
                st.session_state.show_description_editor = False
                st.rerun()
        
        # Gestion des modules
        if st.button("📖 Modifier les modules"):
            st.session_state.show_editor = not st.session_state.show_editor
            st.rerun()
        
        if st.session_state.show_editor:
            selected_module = st.selectbox("Module à modifier", MODULES)
            current_content = charger_contenu_module(selected_module)
            new_content = st.text_area("Contenu du module", value=current_content, height=400)
            
            if st.button("💾 Sauvegarder le module", type="primary"):
                sauvegarder_contenu_module(selected_module, new_content)
                st.success(f"✅ {selected_module} mis à jour !")

# Page Accueil
elif st.session_state.menu_page == "accueil":
    st.markdown(f'<div class="card"><h1 style="text-align: center;">{config["site_title"]}</h1></div>', unsafe_allow_html=True)
    
    if config.get("site_image"):
        try:
            image = Image.open(config["site_image"])
            st.image(image, use_container_width=True, caption="Formation Python - Géologie & Mines")
        except:
            pass
    
    st.markdown(f'<div class="card">{config["site_description"]}</div>', unsafe_allow_html=True)
    
    st.markdown("""
    <div class="card" style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white;">
        <h2 style="text-align: center;">🚀 Prêt(e) à commencer ?</h2>
        <p style="text-align: center; font-size: 1.2rem;">
            Rejoignez des professionnels qui ont transformé leur carrière avec Python !
        </p>
    </div>
    """, unsafe_allow_html=True)

# Page Contenu Formation
elif st.session_state.menu_page == "contenu":
    st.markdown('<div class="card"><h2 class="section-title">📚 Contenu de la Formation</h2></div>', unsafe_allow_html=True)
    
    # Si aucun module n'est sélectionné, sélectionner le premier par défaut
    if not hasattr(st.session_state, 'selected_module') or not st.session_state.selected_module:
        st.session_state.selected_module = MODULES[0].split(" - ")[0]
    
    # Trouver le module complet correspondant à la sélection
    full_module_name = next((m for m in MODULES if m.startswith(st.session_state.selected_module)), MODULES[0])
    
    # Afficher le contenu du module sélectionné
    st.markdown(f'<div class="card"><h3>📖 {full_module_name}</h3></div>', unsafe_allow_html=True)
    content = charger_contenu_module(full_module_name)
    st.markdown(f'<div class="card">{content}</div>', unsafe_allow_html=True)
    
    # Navigation entre modules
    current_index = MODULES.index(full_module_name)
    cols = st.columns(2)
    
    with cols[0]:
        if current_index > 0:
            prev_module = MODULES[current_index - 1]
            if st.button(f"⬅️ {prev_module.split(' - ')[0]}", use_container_width=True):
                st.session_state.selected_module = prev_module.split(" - ")[0]
                st.rerun()
    
    with cols[1]:
        if current_index < len(MODULES) - 1:
            next_module = MODULES[current_index + 1]
            if st.button(f"{next_module.split(' - ')[0]} ➡️", use_container_width=True):
                st.session_state.selected_module = next_module.split(" - ")[0]
                st.rerun()
    
    # Onglets pour navigation rapide
    st.markdown('<div class="card"><h4>Navigation rapide</h4></div>', unsafe_allow_html=True)
    tabs = st.tabs([module.split(" - ")[0] for module in MODULES])
    
    for i, tab in enumerate(tabs):
        with tab:
            if st.button(f"Aller à {MODULES[i].split(' - ')[0]}", use_container_width=True):
                st.session_state.selected_module = MODULES[i].split(" - ")[0]
                st.rerun()

# Page Inscription
elif st.session_state.menu_page == "inscription":
    st.markdown('<div class="card"><h2 class="section-title">📝 Formulaire d\'Inscription</h2></div>', unsafe_allow_html=True)
    
    with st.form("inscription_form", clear_on_submit=True):
        st.markdown('<div class="card"><h3>👤 Informations personnelles</h3></div>', unsafe_allow_html=True)
        
        cols = st.columns(2)
        with cols[0]:
            nom = st.text_input("Nom *")
            prenom = st.text_input("Prénom *")
            cnib = st.text_input("Numéro CNIB *")
        
        with cols[1]:
            telephone = st.text_input("Téléphone *")
            sexe = st.selectbox("Sexe *", ["", "Homme", "Femme"])
            age = st.number_input("Âge *", min_value=16, max_value=80, value=25)
        
        st.markdown('<div class="card"><h3>📅 Préférences de formation</h3></div>', unsafe_allow_html=True)
        
        cols = st.columns(2)
        with cols[0]:
            structure = st.text_input("Structure")
            niveau = st.selectbox("Niveau Python *", ["", "Débutant", "Intermédiaire", "Avancé"])
        
        with cols[1]:
            periode = st.selectbox("Période souhaitée *", ["", "Janvier 2025", "Février 2025", "Mars 2025"])
            option_suivi = st.selectbox("Mode de suivi *", ["", "Présentiel", "En ligne", "Hybride"])
        
        if st.form_submit_button("🚀 S'inscrire", type="primary"):
            errors = []
            
            if not nom or not valider_nom(nom):
                errors.append("❌ Nom invalide")
            if not prenom or not valider_nom(prenom):
                errors.append("❌ Prénom invalide")
            if not cnib or not valider_cnib(cnib):
                errors.append("❌ CNIB invalide")
            if not telephone or not valider_telephone(telephone):
                errors.append("❌ Téléphone invalide")
            if not sexe:
                errors.append("❌ Sexe requis")
            if not niveau:
                errors.append("❌ Niveau requis")
            if not periode:
                errors.append("❌ Période requise")
            if not option_suivi:
                errors.append("❌ Mode de suivi requis")
            
            if errors:
                for error in errors:
                    st.error(error)
            else:
                data = {
                    "Nom": nom.strip().title(),
                    "Prénom": prenom.strip().title(),
                    "Numéro CNIB": cnib.upper().strip(),
                    "Téléphone": telephone.strip(),
                    "Structure": structure.strip() if structure else "Non renseignée",
                    "Période souhaitée": periode,
                    "Sexe": sexe,
                    "Âge": age,
                    "Niveau": niveau,
                    "Option de suivi": option_suivi
                }
                
                success, message = sauvegarder_inscription(data)
                if success:
                    st.success(message)
                    st.balloons()
                else:
                    st.error(message)

# Page Statistiques
elif st.session_state.menu_page == "statistiques":
    if not st.session_state.admin_logged_in:
        st.warning("🔒 Accès réservé aux administrateurs")
        st.session_state.menu_page = "accueil"
        st.rerun()
    else:
        st.markdown('<div class="card"><h2 class="section-title">📊 Statistiques</h2></div>', unsafe_allow_html=True)
        
        df = charger_inscriptions()
        
        if df.empty:
            st.warning("📭 Aucune donnée disponible")
        else:
            cols = st.columns(4)
            with cols[0]:
                st.metric("Total inscriptions", len(df))
            with cols[1]:
                st.metric("Hommes", len(df[df['Sexe'] == 'Homme']))
            with cols[2]:
                st.metric("Femmes", len(df[df['Sexe'] == 'Femme']))
            with cols[3]:
                st.metric("Âge moyen", round(df['Âge'].mean(), 1))
            
            # Graphiques
        col_left, col_right = st.columns(2)
        
        with col_left:
            # Graphique répartition par sexe
            # Graphique répartition par sexe
            if 'Sexe' in df.columns:
                fig_sexe = px.pie(
                    df, 
                    names='Sexe', 
                    title="👥 Répartition par sexe",
                    color_discrete_sequence=['#667eea', '#764ba2']
                    )
                st.plotly_chart(fig_sexe, use_container_width=True)
                # Graphique répartition par niveau
            if 'Niveau' in df.columns:
                niveau_counts = df['Niveau'].value_counts().reset_index()
                niveau_counts.columns = ['Niveau', 'count']  # Renommer les colonnes
                fig_niveau = px.bar(
                     niveau_counts, 
                     x='Niveau', 
                     y='count',
                     title="📊 Répartition par niveau Python",
                     color_discrete_sequence=['#667eea']
                     )
                fig_niveau.update_xaxes(title="Niveau")
                fig_niveau.update_yaxes(title="Nombre d'inscrits")
                st.plotly_chart(fig_niveau, use_container_width=True)
                with col_right:
                    # Graphique répartition par période
                    if 'Période souhaitée' in df.columns:
                        periode_counts = df['Période souhaitée'].value_counts().reset_index()
                        periode_counts.columns = ['Période', 'count']  # Renommer les colonnes
                        fig_periode = px.bar(
                            periode_counts,
                            x='Période',
                            y='count',
                            title="📅 Périodes préférées",
                            color_discrete_sequence=['#764ba2']
                            )
                        fig_periode.update_xaxes(title="Période")
                        fig_periode.update_yaxes(title="Nombre d'inscrits")
                        st.plotly_chart(fig_periode, use_container_width=True)
                        # Graphique répartition par mode de suivi
                        if 'Option de suivi' in df.columns:
                            fig_suivi = px.pie(
                                df, 
                                names='Option de suivi', 
                                title="💻 Modes de suivi préférés",
                                color_discrete_sequence=['#f093fb', '#f5576c', '#4facfe']
                                )
                            st.plotly_chart(fig_suivi, use_container_width=True)

                            # Histogramme des âges
                        if 'Âge' in df.columns:
                             st.markdown("### 📈 Distribution des âges")
                             fig_age = px.histogram(
                                 df, 
                                x='Âge', 
                                nbins=20, 
                                title="Répartition par tranches d'âge",
                                color_discrete_sequence=['#667eea']
                                )
                             fig_age.update_xaxes(title="Âge")
                             fig_age.update_yaxes(title="Nombre d'inscrits")
                        st.plotly_chart(fig_age, use_container_width=True)

# Footer
st.markdown("""
<div class="card">
    <div style="text-align: center;">
        <h3>🐍 Formation Python - Géologie & Mines</h3>
        <p>© 2025 - Tous droits réservés</p>
        <p>📧 formation@tcg-expertise.com | 📱 +226 25 45 67 67 / ‪+33779185080</p>
    </div>
</div>
""", unsafe_allow_html=True)

st.markdown('</div>', unsafe_allow_html=True)  # Fermeture du main-container
