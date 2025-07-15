import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import ttkbootstrap as tb
from openpyxl import load_workbook, Workbook
import os
import re
from datetime import datetime

# Configuration Admin
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "python2025"
admin_logged_in = False
contenu_path = "contenu_formation.txt"  # Fichier texte pour stocker le contenu

# Crée le fichier Excel si inexistant
def initialiser_excel():
    if not os.path.exists("inscriptions.xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.title = "Inscriptions"
        ws.append(["Nom", "Prénom", "Numéro CNIB", "Téléphone", "Structure", 
                   "Période souhaitée", "Sexe", "Âge", "Niveau", "Option de suivi", "Date d'inscription"])
        wb.save("inscriptions.xlsx")

initialiser_excel()

# Interface principale
app = tb.Window(themename="superhero")
app.title("Plateforme d'inscription - Python Géologie & Mines")
app.geometry("1000x700")
app.resizable(True, True)

# Variables globales
notebook = ttk.Notebook(app)
notebook.pack(padx=10, pady=10, fill="both", expand=True)
admin_check = tk.BooleanVar()
frame_login = None

# Fonctions de validation
def valider_telephone(tel):
    """Valide le format du numéro de téléphone"""
    pattern = r'^(\+226|00226)?\s?[0-9]{8}$'
    return re.match(pattern, tel.replace(' ', '')) is not None

def valider_cnib(cnib):
    """Valide le format du numéro CNIB"""
    pattern = r'^[A-Z]{1,2}[0-9]{6,8}$'
    return re.match(pattern, cnib.upper()) is not None

def valider_age(age):
    """Valide l'âge (doit être entre 16 et 80 ans)"""
    try:
        age_int = int(age)
        return 16 <= age_int <= 80
    except ValueError:
        return False

def valider_nom(nom):
    """Valide le nom (pas de chiffres, minimum 2 caractères)"""
    return len(nom) >= 2 and nom.replace(' ', '').replace('-', '').isalpha()

# ==================== ONGLET 1 : CONTENU DE LA FORMATION ====================
frame1 = tb.Frame(notebook)
notebook.add(frame1, text="📘 Contenu Formation")

def authentifier_admin():
    global admin_logged_in
    username = entry_user.get()
    password = entry_pass.get()

    if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
        admin_logged_in = True
        messagebox.showinfo("Connexion réussie", "Vous êtes connecté en tant qu'Admin.")
        frame_login.destroy()
        btn_upload.config(state="normal")
        btn_modifier.config(state="normal")
        btn_stats.config(state="normal")
        btn_export.config(state="normal")
        # Réinitialiser les champs de connexion
        entry_user.delete(0, tk.END)
        entry_pass.delete(0, tk.END)
        admin_check.set(False)
    else:
        messagebox.showerror("Échec", "Nom d'utilisateur ou mot de passe incorrect.")

def activer_connexion():
    global frame_login, entry_user, entry_pass
    if admin_check.get():
        frame_login = tb.Frame(frame1)
        frame_login.pack(pady=10)
        
        tb.Label(frame_login, text="Nom d'utilisateur :", font=("Arial", 10)).grid(row=0, column=0, padx=5, pady=5, sticky="e")
        tb.Label(frame_login, text="Mot de passe :", font=("Arial", 10)).grid(row=1, column=0, padx=5, pady=5, sticky="e")
        
        entry_user = tb.Entry(frame_login, width=20)
        entry_pass = tb.Entry(frame_login, show="*", width=20)
        entry_user.grid(row=0, column=1, padx=5, pady=5)
        entry_pass.grid(row=1, column=1, padx=5, pady=5)
        
        # Binding pour connexion avec Enter
        entry_pass.bind('<Return>', lambda e: authentifier_admin())
        
        btn_connexion = tb.Button(frame_login, text="Connexion", command=authentifier_admin)
        btn_connexion.grid(row=2, column=1, pady=10)
        
        entry_user.focus()
    else:
        if frame_login:
            frame_login.destroy()

cb_admin = tb.Checkbutton(frame1, text="Je suis Admin", variable=admin_check, command=activer_connexion)
cb_admin.pack(pady=5)

# Zone de texte lecture seule pour afficher le contenu
tb.Label(frame1, text="📄 Contenu de la formation :", font=("Arial", 12, "bold")).pack(pady=(10, 5))
text_contenu = tk.Text(frame1, height=20, width=90, wrap="word", font=("Arial", 10))
text_contenu.pack(padx=10, pady=10, fill="both", expand=True)
text_contenu.config(state="disabled")

# Scrollbar pour le texte
scrollbar = ttk.Scrollbar(text_contenu)
scrollbar.pack(side="right", fill="y")
text_contenu.config(yscrollcommand=scrollbar.set)
scrollbar.config(command=text_contenu.yview)

def charger_contenu():
    if os.path.exists(contenu_path):
        with open(contenu_path, "r", encoding="utf-8") as f:
            contenu = f.read()
        text_contenu.config(state="normal")
        text_contenu.delete("1.0", tk.END)
        text_contenu.insert(tk.END, contenu)
        text_contenu.config(state="disabled")
    else:
        text_contenu.config(state="normal")
        text_contenu.delete("1.0", tk.END)
        text_contenu.insert(tk.END, "Aucun contenu de formation n'a été téléversé pour le moment.")
        text_contenu.config(state="disabled")

def televerser():
    if not admin_logged_in:
        messagebox.showwarning("Accès refusé", "Seul un admin peut téléverser le contenu.")
        return
    fichier = filedialog.askopenfilename(filetypes=[("Fichiers texte", "*.txt"), ("Tous les fichiers", "*.*")])
    if fichier:
        try:
            with open(fichier, "r", encoding="utf-8") as f:
                contenu = f.read()
            with open(contenu_path, "w", encoding="utf-8") as f:
                f.write(contenu)
            messagebox.showinfo("Succès", "Contenu mis à jour avec succès.")
            charger_contenu()
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors du téléversement : {str(e)}")

def modifier_contenu():
    if not admin_logged_in:
        messagebox.showwarning("Accès refusé", "Seul un admin peut modifier le contenu.")
        return
    
    # Fenêtre d'édition
    fenetre_edition = tk.Toplevel(app)
    fenetre_edition.title("Modifier le contenu")
    fenetre_edition.geometry("800x600")
    
    text_editor = tk.Text(fenetre_edition, wrap="word", font=("Arial", 10))
    text_editor.pack(padx=10, pady=10, fill="both", expand=True)
    
    # Charger le contenu actuel
    if os.path.exists(contenu_path):
        with open(contenu_path, "r", encoding="utf-8") as f:
            contenu = f.read()
        text_editor.insert(tk.END, contenu)
    
    def sauvegarder():
        contenu = text_editor.get("1.0", tk.END)
        with open(contenu_path, "w", encoding="utf-8") as f:
            f.write(contenu)
        messagebox.showinfo("Succès", "Contenu sauvegardé avec succès.")
        charger_contenu()
        fenetre_edition.destroy()
    
    btn_sauver = tb.Button(fenetre_edition, text="Sauvegarder", command=sauvegarder)
    btn_sauver.pack(pady=5)

# Frame pour les boutons admin
frame_boutons = tb.Frame(frame1)
frame_boutons.pack(pady=10)

btn_upload = tb.Button(frame_boutons, text="🗂️ Téléverser fichier", command=televerser, state="disabled")
btn_upload.pack(side="left", padx=5)

btn_modifier = tb.Button(frame_boutons, text="✏️ Modifier contenu", command=modifier_contenu, state="disabled")
btn_modifier.pack(side="left", padx=5)

# Charger le contenu initial
charger_contenu()

# ==================== ONGLET 2 : INSCRIPTIONS ====================
frame2 = tb.Frame(notebook)
notebook.add(frame2, text="📝 Inscription")

# Créer un frame avec scrollbar pour le formulaire
canvas = tk.Canvas(frame2)
scrollbar_form = ttk.Scrollbar(frame2, orient="vertical", command=canvas.yview)
scrollable_frame = ttk.Frame(canvas)

scrollable_frame.bind(
    "<Configure>",
    lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
)

canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
canvas.configure(yscrollcommand=scrollbar_form.set)

fields = {}
labels = [
    "Nom", "Prénom", "Numéro CNIB", "Téléphone", "Structure", 
    "Période souhaitée", "Âge"
]

# Titre
tb.Label(scrollable_frame, text="Formulaire d'inscription", font=("Arial", 16, "bold")).grid(row=0, column=0, columnspan=4, pady=20)

# Champs de saisie classiques
for i, label in enumerate(labels):
    tb.Label(scrollable_frame, text=label + " :", font=("Arial", 10)).grid(row=i+1, column=0, padx=10, pady=8, sticky="e")
    entry = tb.Entry(scrollable_frame, width=25, font=("Arial", 10))
    entry.grid(row=i+1, column=1, padx=10, pady=8, sticky="w")
    fields[label] = entry

# Combobox - Sexe
tb.Label(scrollable_frame, text="Sexe :", font=("Arial", 10)).grid(row=1, column=2, padx=10, pady=8, sticky="e")
sexe = ttk.Combobox(scrollable_frame, values=["Homme", "Femme"], state="readonly", width=22)
sexe.grid(row=1, column=3, padx=10, pady=8, sticky="w")
fields["Sexe"] = sexe

# Combobox - Niveau
tb.Label(scrollable_frame, text="Niveau :", font=("Arial", 10)).grid(row=2, column=2, padx=10, pady=8, sticky="e")
niveau = ttk.Combobox(scrollable_frame, values=["Débutant", "Intermédiaire", "Avancé"], state="readonly", width=22)
niveau.grid(row=2, column=3, padx=10, pady=8, sticky="w")
fields["Niveau"] = niveau

# Combobox - Option de suivi
tb.Label(scrollable_frame, text="Option de suivi :", font=("Arial", 10)).grid(row=3, column=2, padx=10, pady=8, sticky="e")
option = ttk.Combobox(scrollable_frame, values=["Présentiel", "En ligne", "Hybride"], state="readonly", width=22)
option.grid(row=3, column=3, padx=10, pady=8, sticky="w")
fields["Option de suivi"] = option

# Instructions
instructions = """
Instructions de remplissage :
• Nom et Prénom : Saisir en lettres uniquement
• Numéro CNIB : Format attendu (ex: B1234567)
• Téléphone : Format international (+226) ou national
• Âge : Entre 16 et 80 ans
• Tous les champs sont obligatoires
"""

tb.Label(scrollable_frame, text=instructions, font=("Arial", 9), justify="left", foreground="gray").grid(row=8, column=0, columnspan=4, pady=10)

# Fonction pour enregistrer avec validation
def enregistrer():
    # Récupérer les valeurs
    valeurs = {}
    for label in labels:
        valeurs[label] = fields[label].get().strip()
    
    valeurs["Sexe"] = fields["Sexe"].get()
    valeurs["Niveau"] = fields["Niveau"].get()
    valeurs["Option de suivi"] = fields["Option de suivi"].get()
    
    # Validation
    erreurs = []
    
    # Vérifier que tous les champs sont remplis
    for key, value in valeurs.items():
        if not value:
            erreurs.append(f"Le champ '{key}' est obligatoire.")
    
    if erreurs:
        messagebox.showwarning("Erreur", "\n".join(erreurs))
        return
    
    # Validations spécifiques
    if not valider_nom(valeurs["Nom"]):
        erreurs.append("Le nom doit contenir uniquement des lettres (minimum 2 caractères).")
    
    if not valider_nom(valeurs["Prénom"]):
        erreurs.append("Le prénom doit contenir uniquement des lettres (minimum 2 caractères).")
    
    if not valider_cnib(valeurs["Numéro CNIB"]):
        erreurs.append("Format CNIB invalide (ex: B1234567).")
    
    if not valider_telephone(valeurs["Téléphone"]):
        erreurs.append("Format téléphone invalide (ex: +226 70123456 ou 70123456).")
    
    if not valider_age(valeurs["Âge"]):
        erreurs.append("L'âge doit être un nombre entre 16 et 80 ans.")
    
    if erreurs:
        messagebox.showwarning("Erreur de validation", "\n".join(erreurs))
        return
    
    # Vérifier les doublons CNIB
    try:
        wb = load_workbook("inscriptions.xlsx")
        ws = wb["Inscriptions"]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2] == valeurs["Numéro CNIB"]:
                messagebox.showwarning("Doublon détecté", "Ce numéro CNIB est déjà enregistré.")
                return
        
        # Sauvegarde dans Excel
        data_row = [valeurs[label] for label in labels] + [valeurs["Sexe"], valeurs["Niveau"], valeurs["Option de suivi"], datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
        ws.append(data_row)
        wb.save("inscriptions.xlsx")
        
        messagebox.showinfo("Succès", "Inscription enregistrée avec succès!")
        
        # Réinitialiser les champs
        for field in fields.values():
            if hasattr(field, 'delete'):
                field.delete(0, tk.END)
            else:
                field.set('')
        
        # Remettre le focus sur le premier champ
        fields["Nom"].focus()
        
    except Exception as e:
        messagebox.showerror("Erreur", f"Erreur lors de l'enregistrement : {str(e)}")

def reinitialiser():
    for field in fields.values():
        if hasattr(field, 'delete'):
            field.delete(0, tk.END)
        else:
            field.set('')
    fields["Nom"].focus()

# Boutons
frame_boutons_form = tb.Frame(scrollable_frame)
frame_boutons_form.grid(row=9, column=0, columnspan=4, pady=20)

btn_envoyer = tb.Button(frame_boutons_form, text="✅ S'inscrire", command=enregistrer)
btn_envoyer.pack(side="left", padx=10)

btn_reset = tb.Button(frame_boutons_form, text="🔄 Réinitialiser", command=reinitialiser)
btn_reset.pack(side="left", padx=10)

# Configurer le canvas
canvas.pack(side="left", fill="both", expand=True)
scrollbar_form.pack(side="right", fill="y")

# ==================== ONGLET 3 : STATISTIQUES ====================
frame3 = tb.Frame(notebook)
notebook.add(frame3, text="📊 Statistiques")

def afficher_stats():
    if not admin_logged_in:
        messagebox.showwarning("Accès refusé", "Vous devez être connecté en tant qu'admin pour voir les statistiques.")
        return
    
    try:
        wb = load_workbook("inscriptions.xlsx")
        ws = wb["Inscriptions"]
        data = list(ws.iter_rows(values_only=True))[1:]  # Ignorer l'en-tête
        
        if not data:
            txt_result.delete("1.0", tk.END)
            txt_result.insert(tk.END, "Aucune inscription trouvée.")
            return
        
        total = len(data)
        sexe = {"Homme": 0, "Femme": 0}
        niveau = {"Débutant": 0, "Intermédiaire": 0, "Avancé": 0}
        option = {"Présentiel": 0, "En ligne": 0, "Hybride": 0}
        structures = {}
        ages = []
        
        for row in data:
            # Sexe
            if row[6] in sexe:
                sexe[row[6]] += 1
            
            # Niveau
            if row[8] in niveau:
                niveau[row[8]] += 1
            
            # Option de suivi
            if row[9] in option:
                option[row[9]] += 1
            
            # Structures
            structure = row[4] if row[4] else "Non spécifié"
            structures[structure] = structures.get(structure, 0) + 1
            
            # Âges
            try:
                age = int(row[7])
                ages.append(age)
            except (ValueError, TypeError):
                pass
        
        # Calculs statistiques
        age_moyen = sum(ages) / len(ages) if ages else 0
        age_min = min(ages) if ages else 0
        age_max = max(ages) if ages else 0
        
        # Formatage des résultats
        result = f"📊 STATISTIQUES D'INSCRIPTION\n"
        result += f"{'='*50}\n\n"
        result += f"👥 Nombre total d'inscrits : {total}\n\n"
        
        result += f"🧑 Répartition par sexe :\n"
        for k, v in sexe.items():
            pourcentage = (v/total)*100 if total > 0 else 0
            result += f"   • {k} : {v} ({pourcentage:.1f}%)\n"
        
        result += f"\n📚 Répartition par niveau :\n"
        for k, v in niveau.items():
            pourcentage = (v/total)*100 if total > 0 else 0
            result += f"   • {k} : {v} ({pourcentage:.1f}%)\n"
        
        result += f"\n🌐 Répartition par option de suivi :\n"
        for k, v in option.items():
            pourcentage = (v/total)*100 if total > 0 else 0
            result += f"   • {k} : {v} ({pourcentage:.1f}%)\n"
        
        result += f"\n🏢 Top 5 des structures :\n"
        top_structures = sorted(structures.items(), key=lambda x: x[1], reverse=True)[:5]
        for structure, count in top_structures:
            result += f"   • {structure} : {count}\n"
        
        result += f"\n📈 Statistiques d'âge :\n"
        result += f"   • Âge moyen : {age_moyen:.1f} ans\n"
        result += f"   • Âge minimum : {age_min} ans\n"
        result += f"   • Âge maximum : {age_max} ans\n"
        
        txt_result.delete("1.0", tk.END)
        txt_result.insert(tk.END, result)
        
    except Exception as e:
        messagebox.showerror("Erreur", f"Erreur lors du calcul des statistiques : {str(e)}")

def exporter_donnees():
    if not admin_logged_in:
        messagebox.showwarning("Accès refusé", "Vous devez être connecté en tant qu'admin pour exporter les données.")
        return
    
    try:
        fichier = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if fichier:
            wb = load_workbook("inscriptions.xlsx")
            wb.save(fichier)
            messagebox.showinfo("Succès", f"Données exportées vers : {fichier}")
    
    except Exception as e:
        messagebox.showerror("Erreur", f"Erreur lors de l'exportation : {str(e)}")

# Interface statistiques
frame_stats_buttons = tb.Frame(frame3)
frame_stats_buttons.pack(pady=10)

btn_stats = tb.Button(frame_stats_buttons, text="📊 Afficher les statistiques", command=afficher_stats, state="disabled")
btn_stats.pack(side="left", padx=10)

btn_export = tb.Button(frame_stats_buttons, text="📤 Exporter les données", command=exporter_donnees, state="disabled")
btn_export.pack(side="left", padx=10)

txt_result = tk.Text(frame3, height=25, width=80, font=("Courier", 10))
txt_result.pack(padx=10, pady=10, fill="both", expand=True)

# Scrollbar pour les statistiques
scrollbar_stats = ttk.Scrollbar(txt_result)
scrollbar_stats.pack(side="right", fill="y")
txt_result.config(yscrollcommand=scrollbar_stats.set)
scrollbar_stats.config(command=txt_result.yview)

# Message initial
txt_result.insert(tk.END, "Connectez-vous en tant qu'admin pour voir les statistiques d'inscription.")

# Lancer l'application
if __name__ == "__main__":
    app.mainloop()